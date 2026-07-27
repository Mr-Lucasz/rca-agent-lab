from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from itertools import islice
from pathlib import Path
from typing import Any, Iterable, Protocol

from openpyxl import load_workbook

Record = dict[str, Any]


class InputError(ValueError):
    """Raised when an input file cannot provide valid bug records."""


class RecordReader(Protocol):
    """Extension point for a supported input format."""

    suffixes: frozenset[str]

    def read(self, path: Path) -> list[Record]:
        """Read records from ``path`` without applying domain normalization."""


@dataclass(frozen=True)
class CsvRecordReader:
    suffixes = frozenset({".csv"})

    def read(self, path: Path) -> list[Record]:
        text, encoding = _decode_text(path.read_bytes())
        delimiter = _detect_delimiter(text)
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        header_index = _best_header_index(rows)
        if header_index is None:
            return []
        headers = _deduplicate_headers(rows[header_index])
        records = _records_from_rows(
            rows[header_index + 1 :],
            headers,
            header_index + 2,
        )
        _add_profile(
            records,
            encoding=encoding,
            delimiter=repr(delimiter),
            header_row=header_index + 1,
        )
        return records


@dataclass(frozen=True)
class JsonRecordReader:
    suffixes = frozenset({".json"})

    def read(self, path: Path) -> list[Record]:
        text, encoding = _decode_text(path.read_bytes())
        value = json.loads(text)
        records = _find_record_list(value)
        if records is None:
            raise InputError(
                "JSON deve conter um bug ou uma coleção de objetos de bug."
            )
        flattened = [
            {**_flatten_record(item), "__rca_source_row__": index}
            for index, item in enumerate(records, start=1)
        ]
        _add_profile(flattened, encoding=encoding, container="nested_json")
        return flattened


@dataclass(frozen=True)
class XlsxRecordReader:
    suffixes = frozenset({".xlsx", ".xlsm"})

    def read(self, path: Path) -> list[Record]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            selection = _select_worksheet_header(workbook)
            if selection is None:
                return []
            worksheet, header_index, header_row = selection
            headers = _deduplicate_headers(header_row)
            rows = worksheet.iter_rows(
                min_row=header_index + 2,
                values_only=True,
            )
            records = _records_from_rows(rows, headers, header_index + 2)
            _add_profile(
                records,
                worksheet=worksheet.title,
                header_row=header_index + 1,
            )
            return records
        finally:
            workbook.close()


DEFAULT_READERS: tuple[RecordReader, ...] = (
    CsvRecordReader(),
    JsonRecordReader(),
    XlsxRecordReader(),
)


class BugFileIngestor:
    """Selects a reader by file extension and enforces input invariants."""

    def __init__(self, readers: Iterable[RecordReader] = DEFAULT_READERS) -> None:
        self._readers: dict[str, RecordReader] = {}
        for reader in readers:
            for suffix in reader.suffixes:
                normalized_suffix = suffix.casefold()
                if normalized_suffix in self._readers:
                    raise ValueError(
                        f"Leitor duplicado para a extensão {normalized_suffix}."
                    )
                self._readers[normalized_suffix] = reader

    def ingest(self, path: str | Path) -> list[Record]:
        source = Path(path).expanduser().resolve()
        if not source.exists() or not source.is_file():
            raise InputError(f"Arquivo não encontrado: {source}")
        reader = self._readers.get(source.suffix.casefold())
        if reader is None:
            supported = ", ".join(sorted(self._readers))
            raise InputError(
                f"Formato não suportado. Extensões aceitas: {supported}."
            )
        records = reader.read(source)
        if not records:
            raise InputError("A entrada não contém bugs.")
        return records


_DEFAULT_INGESTOR = BugFileIngestor()


def ingest(path: str | Path) -> list[Record]:
    """Backward-compatible facade for the default file ingestor."""

    return _DEFAULT_INGESTOR.ingest(path)


def _decode_text(payload: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            return payload.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise InputError("Não foi possível detectar a codificação do arquivo.")


def _add_profile(records: list[Record], **profile: Any) -> None:
    for record in records:
        for key, value in profile.items():
            record[f"__rca_{key}__"] = value


def _detect_delimiter(text: str) -> str:
    lines = [line for line in text.splitlines()[:20] if line.strip()]
    scores = {
        delimiter: sum(line.count(delimiter) for line in lines)
        for delimiter in (",", ";", "\t", "|")
    }
    delimiter, score = max(scores.items(), key=lambda item: item[1])
    if score:
        return delimiter
    try:
        return csv.Sniffer().sniff(text[:8192]).delimiter
    except csv.Error:
        return ","


def _best_header_index(rows: list[list[Any]]) -> int | None:
    candidates = []
    for index, row in enumerate(rows[:20]):
        values = [str(value).strip() for value in row if str(value).strip()]
        if len(values) < 2:
            continue
        unique_ratio = len(set(values)) / len(values)
        text_ratio = sum(
            any(char.isalpha() for char in value) for value in values
        ) / len(values)
        candidates.append((len(values) + unique_ratio + text_ratio, -index, index))
    return max(candidates)[2] if candidates else None


def _deduplicate_headers(row: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for value in row:
        header = str(value).strip() if value is not None else ""
        if not header:
            headers.append("")
            continue
        counts[header] = counts.get(header, 0) + 1
        suffix = f"_{counts[header]}" if counts[header] > 1 else ""
        headers.append(f"{header}{suffix}")
    return headers


def _records_from_rows(
    rows: Iterable[Iterable[Any]],
    headers: list[str],
    first_row: int,
) -> list[Record]:
    records = []
    for row_index, row in enumerate(rows, start=first_row):
        values = list(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        record = {
            headers[index]: value
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        record["__rca_source_row__"] = row_index
        records.append(record)
    return records


def _find_record_list(value: Any) -> list[dict[str, Any]] | None:
    if isinstance(value, list) and all(isinstance(item, dict) for item in value):
        return value
    if not isinstance(value, dict):
        return None
    preferred = ("bugs", "defects", "issues", "items", "records", "data", "results")
    for key in preferred:
        found = _find_record_list(value.get(key))
        if found is not None:
            return found
    for nested in value.values():
        found = _find_record_list(nested)
        if found is not None:
            return found
    return [value]


def _flatten_record(value: dict[str, Any], prefix: str = "") -> dict[str, Any]:
    flattened: dict[str, Any] = {}
    for key, item in value.items():
        field = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(item, dict):
            flattened.update(_flatten_record(item, field))
        else:
            flattened[field] = item
    return flattened


def _select_worksheet_header(workbook: Any) -> tuple[Any, int, tuple[Any, ...]] | None:
    candidates = []
    for worksheet in workbook.worksheets:
        preview = list(islice(worksheet.iter_rows(values_only=True), 20))
        header_index = _best_header_index([list(row) for row in preview])
        if header_index is None:
            continue
        width = sum(
            bool(value is not None and str(value).strip())
            for value in preview[header_index]
        )
        candidates.append(
            (
                width,
                min(worksheet.max_row or 0, 1000),
                -header_index,
                worksheet,
                header_index,
                preview[header_index],
            )
        )
    if not candidates:
        return None
    selected = max(candidates, key=lambda item: item[:3])
    return selected[3], selected[4], selected[5]
