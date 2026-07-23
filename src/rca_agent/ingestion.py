from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class InputError(ValueError):
    pass


def _read_csv(path: Path) -> list[dict[str, Any]]:
    sample = path.read_text(encoding="utf-8-sig")[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle, dialect=dialect)]


def _read_json(path: Path) -> list[dict[str, Any]]:
    value = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(value, dict) and isinstance(value.get("bugs"), list):
        value = value["bugs"]
    elif isinstance(value, dict):
        value = [value]
    if not isinstance(value, list) or not all(isinstance(item, dict) for item in value):
        raise InputError("JSON deve ser um bug, uma lista de bugs ou {'bugs': [...]}.")
    return value


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    if not any(headers):
        raise InputError("A primeira linha do XLSX não contém cabeçalhos.")
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        for row in rows
        if any(value is not None and str(value).strip() for value in row)
    ]


def ingest(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path).expanduser().resolve()
    if not source.exists() or not source.is_file():
        raise InputError(f"Arquivo não encontrado: {source}")
    suffix = source.suffix.casefold()
    if suffix == ".csv":
        records = _read_csv(source)
    elif suffix == ".json":
        records = _read_json(source)
    elif suffix in {".xlsx", ".xlsm"}:
        records = _read_xlsx(source)
    else:
        raise InputError("Formato não suportado. Use CSV, XLSX, XLSM ou JSON.")
    if not records:
        raise InputError("A entrada não contém bugs.")
    return records

