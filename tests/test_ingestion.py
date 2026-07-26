import csv
import json
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook

from rca_agent.ingestion import BugFileIngestor, ingest

ROOT = Path(__file__).resolve().parents[1]


def test_ingests_csv_json_and_xlsx():
    tmp_path = ROOT / "reports" / "test-artifacts" / uuid4().hex
    tmp_path.mkdir(parents=True)
    records = [{"id": "BUG-1", "title": "Falha"}, {"id": "BUG-2", "title": "Erro"}]

    csv_path = tmp_path / "bugs.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["id", "title"])
        writer.writeheader()
        writer.writerows(records)

    json_path = tmp_path / "bugs.json"
    json_path.write_text(json.dumps({"bugs": records}), encoding="utf-8")

    xlsx_path = tmp_path / "bugs.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["id", "title"])
    sheet.append(["BUG-1", "Falha"])
    sheet.append(["BUG-2", "Erro"])
    workbook.save(xlsx_path)

    assert len(ingest(csv_path)) == 2
    assert len(ingest(json_path)) == 2
    assert len(ingest(xlsx_path)) == 2


def test_ingestor_accepts_new_reader_without_changing_dispatch_code():
    class TextReader:
        suffixes = frozenset({".bugs"})

        def read(self, path: Path) -> list[dict[str, str]]:
            return [{"id": line} for line in path.read_text(encoding="utf-8").splitlines()]

    tmp_path = ROOT / "reports" / "test-artifacts" / uuid4().hex
    tmp_path.mkdir(parents=True)
    source = tmp_path / "ids.bugs"
    source.write_text("BUG-1\nBUG-2", encoding="utf-8")

    records = BugFileIngestor([TextReader()]).ingest(source)

    assert records == [{"id": "BUG-1"}, {"id": "BUG-2"}]


def test_csv_detects_cp1252_preamble_delimiter_and_real_source_row():
    tmp_path = ROOT / "reports" / "test-artifacts" / uuid4().hex
    tmp_path.mkdir(parents=True)
    source = tmp_path / "empresa-latina.csv"
    source.write_bytes(
        (
            "Relatório interno de defeitos\n"
            "Código;Resumo;Descrição\n"
            "INC-10;Falha de sessão;Usuário não consegue autenticar\n"
        ).encode("cp1252")
    )

    records = ingest(source)

    assert records[0]["Código"] == "INC-10"
    assert records[0]["Resumo"] == "Falha de sessão"
    assert records[0]["__rca_source_row__"] == 3
    assert records[0]["__rca_encoding__"] == "cp1252"
    assert records[0]["__rca_header_row__"] == 2


def test_json_finds_nested_collection_and_flattens_company_structure():
    tmp_path = ROOT / "reports" / "test-artifacts" / uuid4().hex
    tmp_path.mkdir(parents=True)
    source = tmp_path / "nested.json"
    source.write_text(
        json.dumps(
            {
                "export": {
                    "issues": [
                        {
                            "defect": {"id": "CASE-9"},
                            "details": {"summary": "Falha no cadastro"},
                        }
                    ]
                }
            }
        ),
        encoding="utf-8",
    )

    records = ingest(source)

    assert records[0]["defect.id"] == "CASE-9"
    assert records[0]["details.summary"] == "Falha no cadastro"
    assert records[0]["__rca_container__"] == "nested_json"


def test_xlsx_selects_data_sheet_and_header_after_preamble():
    tmp_path = ROOT / "reports" / "test-artifacts" / uuid4().hex
    tmp_path.mkdir(parents=True)
    source = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    workbook.active.append(["Capa do relatório"])
    sheet = workbook.create_sheet("Defects")
    sheet.append(["Exportado em 2026"])
    sheet.append(["Issue Key", "Short Summary", "Impact Level"])
    sheet.append(["ISS-1", "Falha de pagamento", "High"])
    workbook.save(source)

    records = ingest(source)

    assert records[0]["Issue Key"] == "ISS-1"
    assert records[0]["Short Summary"] == "Falha de pagamento"
    assert records[0]["__rca_source_row__"] == 3
    assert records[0]["__rca_worksheet__"] == "Defects"
